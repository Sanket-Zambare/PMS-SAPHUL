"""
Excel import endpoints for bulk creating tasks and projects.
Expected columns:
  Tasks:   title | description | project_name | assigned_to_email | start_date | due_date | status
  Projects: name | description | methodology | start_date | end_date | status
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from datetime import date
import io
import openpyxl

from app.core.security import get_db, get_current_user, require_permission
from app.core.permissions import TASK_CREATE, PROJECT_CREATE
from app.models.task import Task, TaskStatus
from app.models.project import Project, ProjectStatus, ProjectMethodology
from app.models.user import User

router = APIRouter(prefix="/import", tags=["Import"])


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@router.post("/tasks")
async def import_tasks(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission(TASK_CREATE)),
):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    # Normalize headers
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

    created = 0
    failed = []

    for i, row in enumerate(rows[1:], start=2):
        data = dict(zip(headers, row))
        title = str(data.get("title") or "").strip()
        if not title:
            failed.append({"row": i, "error": "Missing title"})
            continue

        project_name = str(data.get("project_name") or "").strip()
        project = db.query(Project).filter(
            Project.name.ilike(project_name), Project.is_deleted == False
        ).first()
        if not project:
            failed.append({"row": i, "error": f"Project '{project_name}' not found"})
            continue

        assigned_to = None
        email = str(data.get("assigned_to_email") or "").strip()
        if email:
            u = db.query(User).filter(User.email.ilike(email)).first()
            if u:
                assigned_to = u.id

        raw_status = str(data.get("status") or "TODO").strip().upper().replace(" ", "_")
        try:
            task_status = TaskStatus[raw_status]
        except KeyError:
            task_status = TaskStatus.TODO

        # Skip if task with same title already exists in this project
        existing = db.query(Task).filter(
            Task.title.ilike(title),
            Task.project_id == project.id,
            Task.is_deleted == False,
        ).first()
        if existing:
            failed.append({"row": i, "error": f"Task '{title}' already exists in '{project_name}' (skipped)"})
            continue

        task = Task(
            title=title,
            description=str(data.get("description") or "").strip() or None,
            project_id=project.id,
            assigned_to=assigned_to,
            created_by=current_user.id,
            status=task_status,
            start_date=parse_date(data.get("start_date")),
            due_date=parse_date(data.get("due_date")),
        )
        db.add(task)
        created += 1

    db.commit()
    return {"created": created, "failed": failed}


@router.post("/projects")
async def import_projects(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_permission(PROJECT_CREATE)),
):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

    created = 0
    failed = []

    for i, row in enumerate(rows[1:], start=2):
        data = dict(zip(headers, row))
        name = str(data.get("name") or "").strip()
        if not name:
            failed.append({"row": i, "error": "Missing name"})
            continue

        raw_method = str(data.get("methodology") or "WATERFALL").strip().upper()
        try:
            methodology = ProjectMethodology[raw_method]
        except KeyError:
            methodology = ProjectMethodology.WATERFALL

        raw_status = str(data.get("status") or "PENDING").strip().upper().replace(" ", "_")
        try:
            proj_status = ProjectStatus[raw_status]
        except KeyError:
            proj_status = ProjectStatus.PENDING

        # Skip if project with same name already exists
        existing = db.query(Project).filter(
            Project.name.ilike(name),
            Project.is_deleted == False,
        ).first()
        if existing:
            failed.append({"row": i, "error": f"Project '{name}' already exists (skipped)"})
            continue

        project = Project(
            name=name,
            description=str(data.get("description") or "").strip() or None,
            methodology=methodology,
            status=proj_status,
            start_date=parse_date(data.get("start_date")),
            end_date=parse_date(data.get("end_date")),
            created_by=current_user.id,
        )
        db.add(project)
        created += 1

    db.commit()
    return {"created": created, "failed": failed}


@router.get("/template/tasks")
async def download_tasks_template():
    """Returns info about the expected Excel columns for task import."""
    return {
        "columns": ["title", "description", "project_name", "assigned_to_email", "start_date", "due_date", "status"],
        "notes": {
            "title": "Required",
            "project_name": "Must match an existing project name exactly",
            "assigned_to_email": "Optional — must match a registered user email",
            "start_date": "Format: YYYY-MM-DD",
            "due_date": "Format: YYYY-MM-DD",
            "status": "TODO | IN_PROGRESS | DONE | BLOCKED (default: TODO)",
        }
    }


@router.get("/template/projects")
async def download_projects_template():
    return {
        "columns": ["name", "description", "methodology", "start_date", "end_date", "status"],
        "notes": {
            "name": "Required",
            "methodology": "WATERFALL | AGILE | HYBRID (default: WATERFALL)",
            "start_date": "Format: YYYY-MM-DD",
            "end_date": "Format: YYYY-MM-DD",
            "status": "PENDING | IN_PROGRESS | COMPLETED | DELAYED (default: PENDING)",
        }
    }
